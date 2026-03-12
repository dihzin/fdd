from __future__ import annotations

import csv
import io
import re
from collections.abc import Iterable
from dataclasses import dataclass

from openpyxl import load_workbook

from app.common.errors.http import bad_request

HEADER_ALIASES = {
    "code": {"code", "req_code", "requirement_code", "requirement_key", "key", "id"},
    "title": {"title", "name", "summary", "requirement_title"},
    "description": {"description", "desc", "details"},
    "requirement_type": {"type", "requirement_type", "category"},
    "module": {"module", "sap_module", "stream", "workstream"},
    "source_system": {"source_system", "source", "source_application", "from_system"},
    "target_system": {"target_system", "target", "target_application", "to_system"},
    "priority": {"priority", "prio", "criticality"},
    "status": {"status", "state"},
}


@dataclass(slots=True)
class ParsedRequirementFile:
    rows: list[dict[str, str | None]]
    mapped_columns: dict[str, str]


class RequirementImportParser:
    def parse(self, filename: str, payload: bytes) -> ParsedRequirementFile:
        extension = filename.rsplit(".", maxsplit=1)[-1].lower() if "." in filename else ""

        if extension == "csv":
            return self._parse_csv(payload)
        if extension in {"xlsx", "xlsm"}:
            return self._parse_excel(payload)

        raise bad_request("Unsupported file type. Use CSV, XLSX or XLSM.")

    def _parse_csv(self, payload: bytes) -> ParsedRequirementFile:
        text = payload.decode("utf-8-sig")
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t") if sample.strip() else csv.excel
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        return self._build_rows(reader.fieldnames or [], reader)

    def _parse_excel(self, payload: bytes) -> ParsedRequirementFile:
        workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return ParsedRequirementFile(rows=[], mapped_columns={})

        normalized_rows = []
        for row in rows:
            normalized_rows.append(
                {
                    str(header): self._clean_cell(value)
                    for header, value in zip(headers, row, strict=False)
                    if header is not None
                }
            )
        return self._build_rows(headers, normalized_rows)

    def _build_rows(
        self,
        headers: Iterable[object],
        raw_rows: Iterable[dict[str, str | None]],
    ) -> ParsedRequirementFile:
        mapped_columns = self._map_headers(headers)
        parsed_rows: list[dict[str, str | None]] = []

        for row in raw_rows:
            mapped_row: dict[str, str | None] = {}
            for source_header, internal_field in mapped_columns.items():
                mapped_row[internal_field] = self._clean_cell(row.get(source_header))
            parsed_rows.append(mapped_row)

        return ParsedRequirementFile(rows=parsed_rows, mapped_columns=mapped_columns)

    def _map_headers(self, headers: Iterable[object]) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for header in headers:
            if header is None:
                continue

            original = str(header)
            normalized = self._normalize_header(original)
            for internal_field, aliases in HEADER_ALIASES.items():
                if normalized in aliases:
                    mapping[original] = internal_field
                    break
        return mapping

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")

    def _clean_cell(self, value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None
